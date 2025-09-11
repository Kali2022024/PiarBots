import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Налаштування логування
logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        
    def create_broadcast_statistics_excel(self, statistics: dict, history: list = None) -> str:
        """Створити Excel файл зі статистикою розсилок"""
        try:
            # Створюємо нову книгу
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Статистика розсилок"
            
            # Налаштування стилів
            self._setup_styles()
            
            # Заголовок
            self._add_header()
            
            # Загальна статистика
            self._add_summary_statistics(statistics['total'])
            
            # Статистика по аккаунтах
            self._add_account_statistics(statistics['by_accounts'])
            
            # Статистика проблемних аккаунтів (FloodWait)
            self._add_floodwait_statistics(statistics.get('floodwait_accounts', []))
            
            # Статистика по чатах
            self._add_chat_statistics(statistics['by_chats'])
            
            # Детальна історія (якщо передана)
            if history:
                self._add_detailed_history(history)
            
            # Автопідгонка ширини колонок
            self._auto_adjust_columns()
            
            # Зберігаємо файл
            filename = f"broadcast_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join("media_files", filename)
            
            # Створюємо папку якщо не існує
            os.makedirs("media_files", exist_ok=True)
            
            self.workbook.save(filepath)
            logger.info(f"✅ Excel файл створено: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"❌ Помилка при створенні Excel файлу: {e}")
            return None
    
    def _setup_styles(self):
        """Налаштування стилів для Excel"""
        # Заголовки
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Підзаголовки
        self.subheader_font = Font(name='Arial', size=11, bold=True, color='000000')
        self.subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Звичайний текст
        self.normal_font = Font(name='Arial', size=10)
        
        # Вирівнювання
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Рамки
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _add_header(self):
        """Додати заголовок"""
        self.worksheet.merge_cells('A1:F1')
        header_cell = self.worksheet['A1']
        header_cell.value = f"📊 Статистика розсилок - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = self.center_alignment
        header_cell.border = self.thin_border
        
        # Пропуск рядка
        self.current_row = 3
    
    def _add_summary_statistics(self, total_stats: dict):
        """Додати загальну статистику"""
        # Заголовок секції
        self.worksheet.merge_cells(f'A{self.current_row}:F{self.current_row}')
        section_cell = self.worksheet[f'A{self.current_row}']
        section_cell.value = "📈 Загальна статистика"
        section_cell.font = self.subheader_font
        section_cell.fill = self.subheader_fill
        section_cell.alignment = self.center_alignment
        section_cell.border = self.thin_border
        self.current_row += 1
        
        # Дані статистики
        stats_data = [
            ["Показник", "Значення"],
            ["Всього відправок", total_stats['total_sends']],
            ["Успішних відправок", total_stats['successful_sends']],
            ["Невдалих відправок", total_stats['failed_sends']],
            ["Унікальних чатів", total_stats['unique_chats']],
            ["Активних аккаунтів", total_stats['unique_accounts']],
            ["Відсоток успішності", f"{(total_stats['successful_sends'] / max(total_stats['total_sends'], 1) * 100):.1f}%"]
        ]
        
        for row_data in stats_data:
            for col_idx, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                cell.value = value
                cell.font = self.normal_font
                cell.alignment = self.left_alignment
                cell.border = self.thin_border
                
                # Підсвічування заголовків
                if col_idx == 1:
                    cell.font = Font(name='Arial', size=10, bold=True)
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            self.current_row += 1
        
        # Пропуск рядка
        self.current_row += 1
    
    def _add_account_statistics(self, account_stats: list):
        """Додати статистику по аккаунтах"""
        if not account_stats:
            return
            
        # Заголовок секції
        self.worksheet.merge_cells(f'A{self.current_row}:F{self.current_row}')
        section_cell = self.worksheet[f'A{self.current_row}']
        section_cell.value = "👤 Статистика по аккаунтах"
        section_cell.font = self.subheader_font
        section_cell.fill = self.subheader_fill
        section_cell.alignment = self.center_alignment
        section_cell.border = self.thin_border
        self.current_row += 1
        
        # Заголовки колонок
        headers = ["Аккаунт", "Ім'я", "Всього", "Успішно", "Невдало", "Успішність"]
        for col_idx, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        self.current_row += 1
        
        # Дані по аккаунтах
        for account_data in account_stats:
            phone, first_name, last_name, total, successful, failed = account_data
            
            # Формуємо ім'я аккаунта
            account_name = f"{first_name or ''} {last_name or ''}".strip()
            if not account_name:
                account_name = phone
            
            # Розраховуємо успішність
            success_rate = (successful / max(total, 1)) * 100
            
            row_data = [phone, account_name, total, successful, failed, f"{success_rate:.1f}%"]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                cell.value = value
                cell.font = self.normal_font
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
                
                # Підсвічування успішності
                if col_idx == 6:  # Колонка успішності
                    if success_rate >= 90:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif success_rate >= 70:
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            self.current_row += 1
        
        # Пропуск рядка
        self.current_row += 1
    
    def _add_floodwait_statistics(self, floodwait_stats: list):
        """Додати статистику проблемних аккаунтів (FloodWait)"""
        if not floodwait_stats:
            return
            
        # Заголовок секції
        self.worksheet.merge_cells(f'A{self.current_row}:D{self.current_row}')
        section_cell = self.worksheet[f'A{self.current_row}']
        section_cell.value = "⚠️ Проблемні аккаунти (FloodWait)"
        section_cell.font = self.subheader_font
        section_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Жовтий фон
        section_cell.alignment = self.center_alignment
        section_cell.border = self.thin_border
        self.current_row += 1
        
        # Заголовки колонок
        headers = ["Номер телефону", "Кількість FloodWait", "Останній FloodWait", "Статус"]
        for col_idx, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        self.current_row += 1
        
        # Дані проблемних аккаунтів
        for account_data in floodwait_stats:
            phone, floodwait_count, last_floodwait = account_data
            
            # Форматуємо дату останнього FloodWait
            if last_floodwait:
                try:
                    if isinstance(last_floodwait, str):
                        last_date = datetime.fromisoformat(last_floodwait.replace('Z', '+00:00')).strftime('%d.%m.%Y %H:%M')
                    else:
                        last_date = str(last_floodwait)
                except:
                    last_date = str(last_floodwait)
            else:
                last_date = "Невідомо"
            
            # Визначаємо статус проблемності
            if floodwait_count >= 10:
                status = "🔴 Критичний"
                status_color = 'FFC7CE'  # Червоний
            elif floodwait_count >= 5:
                status = "🟡 Проблемний"
                status_color = 'FFEB9C'  # Жовтий
            else:
                status = "🟠 Увага"
                status_color = 'FFD8CC'  # Помаранчевий
            
            row_data = [phone, floodwait_count, last_date, status]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                cell.value = value
                cell.font = self.normal_font
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
                
                # Підсвічування статусу
                if col_idx == 4:  # Колонка статусу
                    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
                elif col_idx == 1:  # Колонка номера телефону
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            
            self.current_row += 1
        
        # Пропуск рядка
        self.current_row += 1
    
    def _add_chat_statistics(self, chat_stats: list):
        """Додати статистику по чатах"""
        if not chat_stats:
            return
            
        # Заголовок секції
        self.worksheet.merge_cells(f'A{self.current_row}:F{self.current_row}')
        section_cell = self.worksheet[f'A{self.current_row}']
        section_cell.value = "💬 Топ чатів по кількості відправок"
        section_cell.font = self.subheader_font
        section_cell.fill = self.subheader_fill
        section_cell.alignment = self.center_alignment
        section_cell.border = self.thin_border
        self.current_row += 1
        
        # Заголовки колонок
        headers = ["ID чату", "Назва чату", "Всього", "Успішно", "Невдало", "Успішність"]
        for col_idx, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        self.current_row += 1
        
        # Дані по чатах
        for chat_data in chat_stats:
            chat_id, chat_title, total, successful, failed = chat_data
            
            # Розраховуємо успішність
            success_rate = (successful / max(total, 1)) * 100
            
            row_data = [chat_id, chat_title or "Без назви", total, successful, failed, f"{success_rate:.1f}%"]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                cell.value = value
                cell.font = self.normal_font
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
                
                # Підсвічування успішності
                if col_idx == 6:  # Колонка успішності
                    if success_rate >= 90:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif success_rate >= 70:
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            self.current_row += 1
        
        # Пропуск рядка
        self.current_row += 1
    
    def _add_detailed_history(self, history: list):
        """Додати детальну історію"""
        if not history:
            return
            
        # Заголовок секції
        self.worksheet.merge_cells(f'A{self.current_row}:I{self.current_row}')
        section_cell = self.worksheet[f'A{self.current_row}']
        section_cell.value = "📋 Детальна історія розсилок"
        section_cell.font = self.subheader_font
        section_cell.fill = self.subheader_fill
        section_cell.alignment = self.center_alignment
        section_cell.border = self.thin_border
        self.current_row += 1
        
        # Заголовки колонок
        headers = ["Дата/Час", "Аккаунт", "ID чату", "Назва чату", "Тип повідомлення", "Статус", "Помилка/FloodWait", "Broadcast ID", "Текст повідомлення"]
        for col_idx, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        self.current_row += 1
        
        # Підрахуємо статистику для додаткової інформації
        floodwait_accounts = set()
        successful_chats = set()
        total_messages = len(history)
        successful_messages = 0
        failed_messages = 0
        
        # Дані історії
        for record in history:
            if len(record) < 12:
                continue
                
            # Правильна структура: id, broadcast_id, account_phone, chat_id, chat_title, 
            # message_type, message_text, file_path, file_id, success, error_message, sent_at, ...
            record_id = record[0]
            broadcast_id = record[1]
            account_phone = record[2]
            chat_id = record[3] 
            chat_title = record[4]
            message_type = record[5]
            message_text = record[6] if record[6] else ""
            file_path = record[7] if record[7] else ""
            file_id = record[8] if record[8] else ""
            success = bool(record[9]) if record[9] is not None else False
            error_message = record[10] if record[10] else ""
            sent_at = record[11]
            
            # Підрахунок статистики
            if success:
                successful_messages += 1
            else:
                failed_messages += 1
            
            # Форматуємо дату
            if sent_at:
                try:
                    if isinstance(sent_at, str):
                        date_str = datetime.fromisoformat(sent_at.replace('Z', '+00:00')).strftime('%d.%m.%Y %H:%M')
                    else:
                        date_str = str(sent_at)
                except:
                    date_str = str(sent_at)
            else:
                date_str = "Невідомо"
            
            # Статус та категоризація помилок
            if success:
                status = "✅ Успішно"
                successful_chats.add(str(chat_id))
            else:
                if error_message and isinstance(error_message, str) and "FloodWait" in error_message:
                    status = "⏳ FloodWait"
                    floodwait_accounts.add(str(account_phone))
                else:
                    status = "❌ Помилка"
            
            # Скорочуємо довгий текст повідомлення
            if message_text and isinstance(message_text, str) and len(message_text) > 50:
                short_message = message_text[:50] + "..."
            else:
                short_message = str(message_text) if message_text else ""
            
            row_data = [
                date_str,
                account_phone,
                chat_id,
                chat_title or "Невідома назва",
                message_type or "text",
                status,
                error_message or "",
                broadcast_id,
                short_message
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                cell.value = value
                cell.font = self.normal_font
                cell.alignment = self.left_alignment
                cell.border = self.thin_border
                
                # Підсвічування статусу
                if col_idx == 6:  # Колонка статусу
                    if success:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif "FloodWait" in status:
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                
                # Підсвічування проблемних аккаунтів
                if col_idx == 2 and account_phone in floodwait_accounts:  # Колонка аккаунта
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            
            self.current_row += 1
        
        # Додаємо підсумкову статистику
        self.current_row += 1
        
        # Заголовок підсумку
        self.worksheet.merge_cells(f'A{self.current_row}:I{self.current_row}')
        summary_cell = self.worksheet[f'A{self.current_row}']
        summary_cell.value = "📊 Підсумкова статистика"
        summary_cell.font = self.subheader_font
        summary_cell.fill = self.subheader_fill
        summary_cell.alignment = self.center_alignment
        summary_cell.border = self.thin_border
        self.current_row += 1
        
        # Статистика
        summary_data = [
            ["Показник", "Значення"],
            ["Всього повідомлень", total_messages],
            ["Успішних відправок", successful_messages],
            ["Невдалих відправок", failed_messages],
            ["Унікальних чатів (успішні)", len(successful_chats)],
            ["Проблемних аккаунтів (FloodWait)", len(floodwait_accounts)],
            ["Успішність (%)", f"{(successful_messages / max(total_messages, 1) * 100):.1f}%"]
        ]
        
        if floodwait_accounts:
            summary_data.append(["Аккаунти з FloodWait", ", ".join(sorted(floodwait_accounts))])
        
        for row_data in summary_data:
            for col_idx, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                cell.value = value
                cell.font = self.normal_font
                cell.alignment = self.left_alignment
                cell.border = self.thin_border
                
                # Підсвічування заголовків
                if col_idx == 1:
                    cell.font = Font(name='Arial', size=10, bold=True)
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            self.current_row += 1
    
    def _auto_adjust_columns(self):
        """Автопідгонка ширини колонок"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символів
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
